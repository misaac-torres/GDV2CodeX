from __future__ import annotations

import logging
from pathlib import Path
from typing import Any, Dict, List, Optional

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter

from .config import EXCEL_PATH

logger = logging.getLogger(__name__)

# --- NOMBRES DE HOJAS ---
SHEET_PROYECTOS = "ProyectosTI"
SHEET_DATOS = "Datos"
SHEET_SUG = "Sugerencias"

# --- FILAS / CABECERAS ---
START_ROW_PROYECTOS = 12  # primera fila de datos
HEADER_ROW_PROYECTOS = 11  # fila de cabecera por defecto

# --- COLUMNAS BASE (hoja ProyectosTI) ---
COLS: Dict[str, str] = {
    "ID": "A",
    "Q_RADICADO": "B",
    "PRIORIZADO": "C",
    "ESTADO_PROYECTO": "D",
    "NOMBRE_PROYECTO": "E",
    "DESCRIPCION_PROYECTO": "F",
    "RESPONSABLE_PROYECTO": "G",
    "AREA_SOLICITANTE": "H",
    "FECHA_INICIO": "I",
    "FECHA_ESTIMADA_CIERRE": "J",
    "LINEA_BASE": "K",
    "LINEA_BASE_Q_GESTION": "L",
    "AVANCE": "M",
    "ESTIMADO_AVANCE": "N",
    "PORC_CUMPLIMIENTO": "O",
    "CONTRIBUCION": "P",
    "INICIATIVA_ESTRATEGICA": "Q",
    "TOTAL_DEP": "CN",
    "TOTAL_L": "CO",
    "TOTAL_P": "CP",
    "CUBRIMIENTO_DEP": "CQ",
    "RATING_PO_SYNC": "CR",
}

# --- RANGOS DE DEPENDENCIAS (FLAGS y DESCRIPCIONES) ---
FLAG_START_LETTER = "R"
FLAG_END_LETTER = "BB"
DESC_START_LETTER = "BC"
DESC_END_LETTER = "CM"

FLAG_START_COL = column_index_from_string(FLAG_START_LETTER)
FLAG_END_COL = column_index_from_string(FLAG_END_LETTER)
DESC_START_COL = column_index_from_string(DESC_START_LETTER)
DESC_END_COL = column_index_from_string(DESC_END_LETTER)

DEP_MAPPING: Dict[str, str] = {}
catalogs: Dict[str, List[str]] = {}
CELULA_TREN_MAP: Dict[str, str] = {}


# ======================================================================
# 2. FUNCIONES BÁSICAS PARA EXCEL + UTILIDADES
# ======================================================================


def load_workbook():
    """Abre el libro de Excel (.xlsx)."""
    excel_path = Path(EXCEL_PATH)
    if not excel_path.exists():
        raise FileNotFoundError(
            f"No se encontró el archivo en GD_EXCEL_PATH: {excel_path}."
        )
    try:
        wb = openpyxl.load_workbook(excel_path, keep_vba=False)
    except PermissionError as exc:  # Excel abierto / bloqueado
        raise PermissionError(
            f"No se pudo abrir el archivo {excel_path}. ¿Está abierto o bloqueado?"
        ) from exc
    if SHEET_PROYECTOS not in wb.sheetnames:
        raise KeyError(f"No existe la hoja '{SHEET_PROYECTOS}'.")
    return wb


def get_ws_proyectos(wb=None):
    wb = wb or load_workbook()
    return wb[SHEET_PROYECTOS]


def get_ws_datos(wb=None):
    wb = wb or load_workbook()
    if SHEET_DATOS not in wb.sheetnames:
        raise KeyError(f"No existe la hoja '{SHEET_DATOS}'.")
    return wb[SHEET_DATOS]


def get_ws_sugerencias(wb=None):
    """Devuelve la hoja 'Sugerencias'. Si no existe, la crea con cabecera."""
    wb = wb or load_workbook()
    if SHEET_SUG not in wb.sheetnames:
        ws = wb.create_sheet(SHEET_SUG)
        ws["A1"] = "Usuario"
        ws["B1"] = "Sugerencia"
    else:
        ws = wb[SHEET_SUG]
        if ws.max_row == 1 and ws["A1"].value is None:
            ws["A1"] = "Usuario"
            ws["B1"] = "Sugerencia"
    return ws


def get_header_row_proyectos(ws):
    """Detecta la fila de cabeceras buscando el texto 'ID'."""
    id_col_idx = column_index_from_string(COLS["ID"])
    for r in range(1, START_ROW_PROYECTOS):
        v = ws.cell(row=r, column=id_col_idx).value
        if isinstance(v, str) and v.strip().lower() == "id":
            return r
    return HEADER_ROW_PROYECTOS


def get_unique_list_from_column(ws, col_letter, start_row=2):
    """Devuelve una lista ordenada sin duplicados de una columna."""
    values = set()
    col_idx = column_index_from_string(col_letter)
    for row in range(start_row, ws.max_row + 1):
        value = ws.cell(row=row, column=col_idx).value
        if value not in (None, ""):
            values.add(str(value))
    return sorted(values)


def get_next_row_and_id(ws, id_col_letter="A", start_row=START_ROW_PROYECTOS):
    """Busca la siguiente fila libre y el siguiente ID numérico."""
    id_col_idx = column_index_from_string(id_col_letter)
    max_row_used = 0
    max_id_found = 0

    for row in range(start_row, ws.max_row + 1):
        val = ws.cell(row=row, column=id_col_idx).value
        if val not in (None, ""):
            max_row_used = row
            if isinstance(val, (int, float)):
                max_id_found = max(max_id_found, int(val))

    next_row = start_row if max_row_used == 0 else max_row_used + 1
    next_id = max_id_found + 1 if max_id_found > 0 else 1
    return next_row, next_id


def find_column_by_header(ws, header_name, header_row=1):
    """Busca una columna por el texto exacto del header (case-insensitive)."""
    if not header_name:
        return None
    target = str(header_name).strip().lower()
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col_idx).value
        if val is None:
            continue
        if str(val).strip().lower() == target:
            return col_idx
    return None


def find_column_by_header_in_range(ws, header_name, start_col_idx, end_col_idx, header_row):
    """Busca una columna por header, restringida a un rango de columnas."""
    if not header_name:
        return None
    target = str(header_name).strip().lower()
    for col_idx in range(start_col_idx, end_col_idx + 1):
        val = ws.cell(row=header_row, column=col_idx).value
        if val is None:
            continue
        if str(val).strip().lower() == target:
            return col_idx
    return None


def find_area_tren_coe_col(ws):
    """Intenta localizar la columna de Area/Tren/CoE en la hoja ProyectosTI."""
    header_row = get_header_row_proyectos(ws)
    candidate_idx = None
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col_idx).value
        if not val:
            continue
        s = str(val).strip().lower()
        if "tren" in s and "coe" in s:
            return col_idx
        if s in ("area tren coe", "area/tren/coe"):
            candidate_idx = col_idx
    return candidate_idx


def to_num_cell(v):
    """Convierte cualquier valor de celda a float, tolerando texto, %, comas, etc."""
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s == "":
        return 0.0
    s = s.replace("%", "").replace(",", ".")
    try:
        return float(s)
    except Exception:
        return 0.0


# ======================================================================
# 3. CARGA DE CATÁLOGOS Y MAPPING DE DEPENDENCIAS / TRENES (HOJA DATOS)
# ======================================================================


def load_dependency_mapping(wb=None):
    """Lee mapeo celula → cabecera de descripción."""
    wb = wb or load_workbook()
    ws_d = get_ws_datos(wb)
    col_cel_dep_idx = find_column_by_header(ws_d, "Celula Dependencia", header_row=1)
    col_desc_dep_idx = find_column_by_header(ws_d, "Celula Descripcion Dependencia", header_row=1)
    mapping: Dict[str, str] = {}
    if not col_cel_dep_idx or not col_desc_dep_idx:
        return mapping

    for row in range(2, ws_d.max_row + 1):
        cel_name = ws_d.cell(row=row, column=col_cel_dep_idx).value
        desc_header = ws_d.cell(row=row, column=col_desc_dep_idx).value
        if cel_name and desc_header:
            mapping[str(cel_name).strip()] = str(desc_header).strip()
    return mapping


def load_catalogs():
    """Carga catálogos y mappings desde la hoja Datos."""
    global DEP_MAPPING, catalogs, CELULA_TREN_MAP

    wb = load_workbook()
    ws_d = get_ws_datos(wb)

    estados_list = get_unique_list_from_column(ws_d, "A")
    priorizacion_list = get_unique_list_from_column(ws_d, "B")
    responsables_list = get_unique_list_from_column(ws_d, "C")
    areas_list = get_unique_list_from_column(ws_d, "D")

    CELULA_TREN_MAP = {}
    col_tren_idx = column_index_from_string("E")
    col_cel_idx = column_index_from_string("F")
    for row in range(2, ws_d.max_row + 1):
        tren_val = ws_d.cell(row=row, column=col_tren_idx).value
        cel_val = ws_d.cell(row=row, column=col_cel_idx).value
        if tren_val and cel_val:
            CELULA_TREN_MAP[str(cel_val).strip()] = str(tren_val).strip()

    area_tren_coe_list = sorted({v for v in CELULA_TREN_MAP.values()})
    celulas_dep_list = sorted({k for k in CELULA_TREN_MAP.keys()})

    iniciativas_list: List[str] = []
    col_ini_idx = find_column_by_header(ws_d, "Iniciativa Estrategica", header_row=1)
    if col_ini_idx:
        col_ini_letter = get_column_letter(col_ini_idx)
        iniciativas_list = get_unique_list_from_column(ws_d, col_ini_letter)

    DEP_MAPPING = load_dependency_mapping(wb)
    if DEP_MAPPING:
        celulas_dep_from_mapping = sorted(DEP_MAPPING.keys())
        celulas_dep_list = sorted(set(celulas_dep_list) | set(celulas_dep_from_mapping))

    catalogs = {
        "estados": estados_list,
        "q_rad": priorizacion_list,
        "responsables": responsables_list,
        "areas": areas_list,
        "area_tren_coe": area_tren_coe_list,
        "celulas_dep": celulas_dep_list,
        "iniciativas": iniciativas_list,
    }


try:
    load_catalogs()
except Exception as exc:  # pragma: no cover - fallback solo en entornos sin Excel
    logger.warning("No se pudieron cargar catálogos desde 'Datos': %s", exc)
    catalogs = {
        "estados": ["Nuevo", "En curso", "Detenido", "Cancelado", "Finalizado"],
        "q_rad": ["1Q/25", "2Q/25"],
        "responsables": ["Responsable 1"],
        "areas": ["Área 1"],
        "area_tren_coe": ["Tren X"],
        "celulas_dep": ["Célula A"],
        "iniciativas": ["Inic. 1"],
    }
    DEP_MAPPING = {}
    CELULA_TREN_MAP = {}


# ======================================================================
# 4. NÚCLEO DEPENDENCIAS: AGREGADOS
# ======================================================================


def compute_dep_aggregates(dep_list):
    """Calcula agregados de dependencias (total, L, P, cubrimiento)."""
    flags = [
        (d.get("codigo") or "").strip().upper()
        for d in dep_list
        if (d.get("equipo") or "").strip()
    ]
    flags = [f for f in flags if f in ("L", "P")]
    total_dep = len(flags)
    total_L = sum(1 for f in flags if f == "L")
    total_P = sum(1 for f in flags if f == "P")
    cubrimiento = total_P / total_dep if total_dep > 0 else 0.0
    return total_dep, total_L, total_P, cubrimiento


def write_dep_aggregates(ws, row, dep_list):
    total_dep, total_L, total_P, cub = compute_dep_aggregates(dep_list)
    cn = column_index_from_string(COLS["TOTAL_DEP"])
    co = column_index_from_string(COLS["TOTAL_L"])
    cp = column_index_from_string(COLS["TOTAL_P"])
    cq = column_index_from_string(COLS["CUBRIMIENTO_DEP"])

    ws.cell(row=row, column=cn).value = total_dep
    ws.cell(row=row, column=co).value = total_L
    ws.cell(row=row, column=cp).value = total_P
    ws.cell(row=row, column=cq).value = cub


def apply_dependencies_to_row(ws, row, dep_list):
    """Aplica dependencias dinámicas en la fila `row`."""
    header_row = get_header_row_proyectos(ws)

    for equipo, desc_header in DEP_MAPPING.items():
        flag_col_idx = find_column_by_header_in_range(
            ws, equipo, FLAG_START_COL, FLAG_END_COL, header_row
        )
        if flag_col_idx:
            ws.cell(row=row, column=flag_col_idx).value = None

        if desc_header:
            desc_col_idx = find_column_by_header_in_range(
                ws, desc_header, DESC_START_COL, DESC_END_COL, header_row
            )
            if desc_col_idx:
                ws.cell(row=row, column=desc_col_idx).value = None

    for dep in dep_list:
        equipo = (dep.get("equipo") or "").strip()
        codigo = (dep.get("codigo") or "").strip().upper()
        texto = (dep.get("descripcion") or "").strip()

        if not equipo or codigo not in ("P", "L"):
            continue

        desc_header = DEP_MAPPING.get(equipo)

        flag_col_idx = find_column_by_header_in_range(
            ws, equipo, FLAG_START_COL, FLAG_END_COL, header_row
        )
        if flag_col_idx:
            ws.cell(row=row, column=flag_col_idx).value = codigo

        if desc_header:
            desc_col_idx = find_column_by_header_in_range(
                ws, desc_header, DESC_START_COL, DESC_END_COL, header_row
            )
            if desc_col_idx and texto:
                ws.cell(row=row, column=desc_col_idx).value = texto

    write_dep_aggregates(ws, row, dep_list)


# ======================================================================
# 5. ALTA DE PROYECTOS (NUEVO REGISTRO + DEPENDENCIAS)
# ======================================================================


def write_project_with_dependencies(project, dep_list):
    """Inserta un nuevo proyecto y aplica sus dependencias + métricas."""
    wb = load_workbook()
    ws = get_ws_proyectos(wb)

    next_row, next_id = get_next_row_and_id(
        ws, id_col_letter=COLS["ID"], start_row=START_ROW_PROYECTOS
    )

    project = project.copy()
    project["ID"] = next_id

    for field, col_letter in COLS.items():
        if field in project:
            col_idx = column_index_from_string(col_letter)
            ws.cell(row=next_row, column=col_idx).value = project.get(field)

    apply_dependencies_to_row(ws, next_row, dep_list)

    wb.save(EXCEL_PATH)
    return next_row, next_id


# ======================================================================
# 6. CONSULTAS / RESÚMENES
# ======================================================================


def get_all_project_names():
    wb = load_workbook()
    ws = get_ws_proyectos(wb)
    name_col_idx = column_index_from_string(COLS["NOMBRE_PROYECTO"])
    names = set()
    for row in range(START_ROW_PROYECTOS, ws.max_row + 1):
        val = ws.cell(row=row, column=name_col_idx).value
        if val not in (None, ""):
            names.add(str(val).strip())
    return sorted(names)


def summarize_by_equipo(equipo_name):
    wb = load_workbook()
    ws = get_ws_proyectos(wb)
    header_row = get_header_row_proyectos(ws)

    col_flag_idx = find_column_by_header_in_range(
        ws, equipo_name, FLAG_START_COL, FLAG_END_COL, header_row
    )
    if not col_flag_idx:
        return {"found": False, "msg": f"No se encontró la columna '{equipo_name}' en R:BB."}

    name_col_idx = column_index_from_string(COLS["NOMBRE_PROYECTO"])
    q_col_idx = column_index_from_string(COLS["Q_RADICADO"])

    total = pendientes = negociadas = 0
    rows = []

    for row in range(START_ROW_PROYECTOS, ws.max_row + 1):
        flag = ws.cell(row=row, column=col_flag_idx).value
        if flag is None or str(flag).strip() == "":
            continue
        flag_up = str(flag).strip().upper()
        if flag_up not in ("P", "L"):
            continue

        total += 1
        if flag_up == "P":
            pendientes += 1
        else:
            negociadas += 1

        nombre = ws.cell(row=row, column=name_col_idx).value
        qrad = ws.cell(row=row, column=q_col_idx).value
        rows.append({"fila": row, "Q_RADICADO": qrad, "PROYECTO": nombre, "FLAG": flag_up})

    pct_pend = (pendientes / total * 100) if total > 0 else 0.0
    return {
        "found": True,
        "equipo": equipo_name,
        "total": total,
        "pendientes": pendientes,
        "negociadas": negociadas,
        "pct_pendientes": pct_pend,
        "rows": rows,
    }


def summarize_by_proyecto(nombre_proyecto):
    wb = load_workbook()
    ws = get_ws_proyectos(wb)
    header_row = get_header_row_proyectos(ws)

    name_col_idx = column_index_from_string(COLS["NOMBRE_PROYECTO"])
    q_col_idx = column_index_from_string(COLS["Q_RADICADO"])

    target_row = None
    for row in range(START_ROW_PROYECTOS, ws.max_row + 1):
        val = ws.cell(row=row, column=name_col_idx).value
        if val and str(val).strip() == nombre_proyecto:
            target_row = row
            break

    if not target_row:
        return {"found": False, "msg": f"No se encontró el proyecto '{nombre_proyecto}'."}

    detalles = []

    for equipo, desc_header in DEP_MAPPING.items():
        flag_col_idx = find_column_by_header_in_range(
            ws, equipo, FLAG_START_COL, FLAG_END_COL, header_row
        )
        if not flag_col_idx:
            continue

        flag = ws.cell(row=target_row, column=flag_col_idx).value
        if flag is None or str(flag).strip() == "":
            continue

        flag_up = str(flag).strip().upper()
        if flag_up not in ("P", "L"):
            continue

        desc = ""
        if desc_header:
            desc_col_idx = find_column_by_header_in_range(
                ws, desc_header, DESC_START_COL, DESC_END_COL, header_row
            )
            if desc_col_idx:
                desc = ws.cell(row=target_row, column=desc_col_idx).value or ""

        detalles.append({"equipo": equipo, "FLAG": flag_up, "descripcion": desc})

    total = len(detalles)
    pendientes = sum(1 for d in detalles if d["FLAG"] == "P")
    negociadas = sum(1 for d in detalles if d["FLAG"] == "L")
    pct_pend = (pendientes / total * 100) if total > 0 else 0.0

    lb_col = column_index_from_string(COLS["LINEA_BASE"])
    av_col = column_index_from_string(COLS["AVANCE"])
    est_col = column_index_from_string(COLS["ESTIMADO_AVANCE"])

    lb = to_num_cell(ws.cell(row=target_row, column=lb_col).value)
    av = to_num_cell(ws.cell(row=target_row, column=av_col).value)
    est = to_num_cell(ws.cell(row=target_row, column=est_col).value)
    qrad = ws.cell(row=target_row, column=q_col_idx).value

    cn = column_index_from_string(COLS["TOTAL_DEP"])
    co = column_index_from_string(COLS["TOTAL_L"])
    cp = column_index_from_string(COLS["TOTAL_P"])
    cq = column_index_from_string(COLS["CUBRIMIENTO_DEP"])

    total_dep_xl = to_num_cell(ws.cell(row=target_row, column=cn).value)
    total_L_xl = to_num_cell(ws.cell(row=target_row, column=co).value)
    total_P_xl = to_num_cell(ws.cell(row=target_row, column=cp).value)
    cub_xl = to_num_cell(ws.cell(row=target_row, column=cq).value)

    return {
        "found": True,
        "fila": target_row,
        "proyecto": nombre_proyecto,
        "Q_RADICADO": qrad,
        "total_dep": total,
        "pendientes": pendientes,
        "negociadas": negociadas,
        "pct_pendientes": pct_pend,
        "detalles": detalles,
        "linea_base": float(lb),
        "avance": float(av),
        "estimado": float(est),
        "total_dep_xl": total_dep_xl,
        "total_L_xl": total_L_xl,
        "total_P_xl": total_P_xl,
        "cub_xl": float(cub_xl),
    }


# ======================================================================
# 7. ACTUALIZAR PROYECTO EXISTENTE (AVANCE + DEPENDENCIAS)
# ======================================================================


def update_project_row_and_dependencies(row, avance, estimado, dep_list):
    wb = load_workbook()
    ws = get_ws_proyectos(wb)

    lb_col = column_index_from_string(COLS["LINEA_BASE"])
    av_col = column_index_from_string(COLS["AVANCE"])
    est_col = column_index_from_string(COLS["ESTIMADO_AVANCE"])
    pct_col = column_index_from_string(COLS["PORC_CUMPLIMIENTO"])

    linea_base = to_num_cell(ws.cell(row=row, column=lb_col).value)
    old_av = to_num_cell(ws.cell(row=row, column=av_col).value)
    old_es = to_num_cell(ws.cell(row=row, column=est_col).value)

    new_av = float(avance) if avance is not None else float(old_av)
    new_es = float(estimado) if estimado is not None else float(old_es)

    ws.cell(row=row, column=av_col).value = new_av
    ws.cell(row=row, column=est_col).value = new_es

    pct_cumpl = new_av / new_es if new_es > 0 else 0.0
    ws.cell(row=row, column=pct_col).value = pct_cumpl

    apply_dependencies_to_row(ws, row, dep_list)

    wb.save(EXCEL_PATH)

    var_vs_lb = new_av - float(linea_base)
    return {
        "linea_base": float(linea_base),
        "avance": new_av,
        "estimado": new_es,
        "pct_cumpl": pct_cumpl * 100,
        "var_vs_lb_pp": var_vs_lb * 100,
    }


# ======================================================================
# 8. CÁLCULO DE MÉTRICAS AGREGADAS
# ======================================================================


def compute_metrics(scope="all", filter_value=None):
    wb = load_workbook()
    ws = get_ws_proyectos(wb)
    header_row = get_header_row_proyectos(ws)

    name_col_idx = column_index_from_string(COLS["NOMBRE_PROYECTO"])
    prior_col_idx = column_index_from_string(COLS["PRIORIZADO"])
    av_col_idx = column_index_from_string(COLS["AVANCE"])
    cn_idx = column_index_from_string(COLS["TOTAL_DEP"])
    co_idx = column_index_from_string(COLS["TOTAL_L"])
    cp_idx = column_index_from_string(COLS["TOTAL_P"])

    cel_flag_col_idx = None
    if scope == "celula" and filter_value:
        cel_flag_col_idx = find_column_by_header_in_range(
            ws, filter_value, FLAG_START_COL, FLAG_END_COL, header_row
        )

    total_projects = 0
    total_dep = 0.0
    total_L = 0.0
    total_P = 0.0
    sum_avance = 0.0
    pri_count = 0
    pri_avance_sum = 0.0
    no_pri_count = 0
    no_pri_avance_sum = 0.0

    for row in range(START_ROW_PROYECTOS, ws.max_row + 1):
        name = ws.cell(row=row, column=name_col_idx).value
        if not name:
            continue

        if scope == "area" and filter_value:
            tren_match = False
            for equipo in DEP_MAPPING.keys():
                flag_col_idx = find_column_by_header_in_range(
                    ws, equipo, FLAG_START_COL, FLAG_END_COL, header_row
                )
                if not flag_col_idx:
                    continue
                flag_val = ws.cell(row=row, column=flag_col_idx).value
                if not flag_val:
                    continue
                flag_up = str(flag_val).strip().upper()
                if flag_up not in ("P", "L"):
                    continue
                tren_val = CELULA_TREN_MAP.get(equipo)
                if tren_val and str(tren_val).strip() == str(filter_value).strip():
                    tren_match = True
                    break
            if not tren_match:
                continue

        if scope == "celula" and filter_value:
            if not cel_flag_col_idx:
                continue
            flag_val = ws.cell(row=row, column=cel_flag_col_idx).value
            if not flag_val or str(flag_val).strip().upper() not in ("P", "L"):
                continue

        total_projects += 1

        dep_val = to_num_cell(ws.cell(row=row, column=cn_idx).value)
        L_val = to_num_cell(ws.cell(row=row, column=co_idx).value)
        P_val = to_num_cell(ws.cell(row=row, column=cp_idx).value)

        total_dep += dep_val
        total_L += L_val
        total_P += P_val

        av_val = to_num_cell(ws.cell(row=row, column=av_col_idx).value)
        sum_avance += av_val

        pri_val = ws.cell(row=row, column=prior_col_idx).value
        pri_str = str(pri_val).strip().upper() if pri_val not in (None, "") else ""

        if pri_str == "SI":
            pri_count += 1
            pri_avance_sum += av_val
        else:
            no_pri_count += 1
            no_pri_avance_sum += av_val

    avg_avance = sum_avance / total_projects if total_projects > 0 else 0.0
    avg_pri = pri_avance_sum / pri_count if pri_count > 0 else 0.0
    avg_no_pri = no_pri_avance_sum / no_pri_count if no_pri_count > 0 else 0.0
    cobertura_pct = (total_P / total_dep) * 100.0 if total_dep > 0 else 0.0

    return {
        "total_projects": int(total_projects),
        "total_dep": float(total_dep),
        "total_L": float(total_L),
        "total_P": float(total_P),
        "cobertura_pct": float(cobertura_pct),
        "avg_avance": float(avg_avance),
        "avg_pri": float(avg_pri),
        "avg_no_pri": float(avg_no_pri),
        "num_pri": int(pri_count),
    }


# ======================================================================
# 13. LÓGICA DE MESA DE EXPERTOS / ALISTAMIENTO
# ======================================================================


def _collect_board_projects(tren_filter: Optional[str] = None):
    wb = load_workbook()
    ws = get_ws_proyectos(wb)
    header_row = get_header_row_proyectos(ws)

    id_idx = column_index_from_string(COLS["ID"])
    q_idx = column_index_from_string(COLS["Q_RADICADO"])
    pri_idx = column_index_from_string(COLS["PRIORIZADO"])
    estado_idx = column_index_from_string(COLS["ESTADO_PROYECTO"])
    nom_idx = column_index_from_string(COLS["NOMBRE_PROYECTO"])
    desc_idx = column_index_from_string(COLS["DESCRIPCION_PROYECTO"])
    cont_idx = column_index_from_string(COLS["CONTRIBUCION"])
    inic_idx = column_index_from_string(COLS["INICIATIVA_ESTRATEGICA"])

    total_dep_idx = column_index_from_string(COLS["TOTAL_DEP"])
    total_L_idx = column_index_from_string(COLS["TOTAL_L"])
    total_P_idx = column_index_from_string(COLS["TOTAL_P"])
    cub_idx = column_index_from_string(COLS["CUBRIMIENTO_DEP"])
    rating_idx = column_index_from_string(COLS["RATING_PO_SYNC"])

    tren_filter = tren_filter or None
    if tren_filter:
        equipos_tren = [
            c for c, t in CELULA_TREN_MAP.items() if str(t).strip() == str(tren_filter).strip()
        ]
    else:
        equipos_tren = []

    proyectos = []

    for row in range(START_ROW_PROYECTOS, ws.max_row + 1):
        estado_val = ws.cell(row=row, column=estado_idx).value
        if not estado_val:
            continue
        estado_str = str(estado_val).strip().lower()
        if estado_str not in ("nuevo", "en curso", "en curso "):
            continue

        if tren_filter and equipos_tren:
            match_tren = False
            for cel in equipos_tren:
                flag_col_idx = find_column_by_header_in_range(
                    ws, cel, FLAG_START_COL, FLAG_END_COL, header_row
                )
                if not flag_col_idx:
                    continue
                flag_val = ws.cell(row=row, column=flag_col_idx).value
                if flag_val and str(flag_val).strip().upper() in ("P", "L"):
                    match_tren = True
                    break
            if not match_tren:
                continue

        id_val = ws.cell(row=row, column=id_idx).value
        q_val = ws.cell(row=row, column=q_idx).value
        pri_val = ws.cell(row=row, column=pri_idx).value
        nom_val = ws.cell(row=row, column=nom_idx).value
        desc_val = ws.cell(row=row, column=desc_idx).value
        cont_val = ws.cell(row=row, column=cont_idx).value
        inic_val = ws.cell(row=row, column=inic_idx).value

        total_dep = to_num_cell(ws.cell(row=row, column=total_dep_idx).value)
        total_L = to_num_cell(ws.cell(row=row, column=total_L_idx).value)
        total_P = to_num_cell(ws.cell(row=row, column=total_P_idx).value)
        cub = to_num_cell(ws.cell(row=row, column=cub_idx).value)
        rating_po = to_num_cell(ws.cell(row=row, column=rating_idx).value)

        pendientes = []
        negociadas = []
        for equipo, _desc_header in DEP_MAPPING.items():
            flag_idx = find_column_by_header_in_range(
                ws, equipo, FLAG_START_COL, FLAG_END_COL, header_row
            )
            if not flag_idx:
                continue
            flag_val = ws.cell(row=row, column=flag_idx).value
            if not flag_val:
                continue
            flag_up = str(flag_val).strip().upper()
            if flag_up == "P":
                pendientes.append(equipo)
            elif flag_up == "L":
                negociadas.append(equipo)

        desc_short = ""
        if desc_val:
            s = str(desc_val).strip()
            desc_short = s if len(s) <= 80 else s[:77] + "..."

        proyectos.append(
            {
                "row": row,
                "id": id_val,
                "q_rad": q_val,
                "estado": estado_val,
                "priorizado": (str(pri_val).strip().upper() if pri_val not in (None, "") else ""),
                "nombre": nom_val,
                "descripcion_corta": desc_short,
                "contribucion": to_num_cell(cont_val),
                "inic_estrategica": inic_val or "",
                "total_dep": total_dep,
                "total_L": total_L,
                "total_P": total_P,
                "cub": cub,
                "rating_po": rating_po,
                "pendientes_list": pendientes,
                "negociadas_list": negociadas,
            }
        )

    proyectos.sort(key=lambda x: (str(x["q_rad"]), str(x["nombre"])))
    return proyectos


def _update_po_rating(row: int, rating_value: int):
    wb = load_workbook()
    ws = get_ws_proyectos(wb)
    rating_idx = column_index_from_string(COLS["RATING_PO_SYNC"])
    ws.cell(row=row, column=rating_idx).value = int(rating_value)
    wb.save(EXCEL_PATH)


def _update_expert_prioritization(row: int, priorizado: str):
    wb = load_workbook()
    ws = get_ws_proyectos(wb)
    pri_idx = column_index_from_string(COLS["PRIORIZADO"])
    ws.cell(row=row, column=pri_idx).value = priorizado
    wb.save(EXCEL_PATH)


# ======================================================================
# 16. FEEDBACK – HOJA SUGERENCIAS
# ======================================================================


def _append_suggestion(usuario: str, texto: str):
    wb = load_workbook()
    ws = get_ws_sugerencias(wb)

    next_row = ws.max_row + 1
    if next_row == 2 and ws["A1"].value is None:
        ws["A1"] = "Usuario"
        ws["B1"] = "Sugerencia"
        next_row = 2

    ws.cell(row=next_row, column=1).value = usuario or ""
    ws.cell(row=next_row, column=2).value = texto or ""
    wb.save(EXCEL_PATH)


def _get_last_suggestions(limit: int = 5):
    wb = load_workbook()
    ws = get_ws_sugerencias(wb)

    rows = []
    for row in range(2, ws.max_row + 1):
        usuario = ws.cell(row=row, column=1).value
        texto = ws.cell(row=row, column=2).value
        if usuario is None and texto is None:
            continue
        rows.append(
            {
                "usuario": str(usuario) if usuario is not None else "",
                "texto": str(texto) if texto is not None else "",
            }
        )

    if not rows:
        return []

    return rows[-limit:]


# ======================================================================
# UTILIDADES ADICIONALES PARA EL BACKEND
# ======================================================================


def find_project_row_by_id(project_id: str) -> Optional[int]:
    wb = load_workbook()
    ws = get_ws_proyectos(wb)
    id_col_idx = column_index_from_string(COLS["ID"])
    for row in range(START_ROW_PROYECTOS, ws.max_row + 1):
        val = ws.cell(row=row, column=id_col_idx).value
        if val is None:
            continue
        if str(val).strip() == str(project_id).strip():
            return row
    return None


def read_project_row(ws, row: int) -> Dict[str, Any]:
    data: Dict[str, Any] = {}
    for field, col_letter in COLS.items():
        col_idx = column_index_from_string(col_letter)
        data[field] = ws.cell(row=row, column=col_idx).value
    return data


def get_project_dependencies(ws, row: int) -> List[Dict[str, Any]]:
    header_row = get_header_row_proyectos(ws)
    deps: List[Dict[str, Any]] = []
    for equipo, desc_header in DEP_MAPPING.items():
        flag_col_idx = find_column_by_header_in_range(
            ws, equipo, FLAG_START_COL, FLAG_END_COL, header_row
        )
        if not flag_col_idx:
            continue
        flag = ws.cell(row=row, column=flag_col_idx).value
        if flag is None or str(flag).strip() == "":
            continue
        flag_up = str(flag).strip().upper()
        if flag_up not in ("P", "L"):
            continue
        desc_val = ""
        if desc_header:
            desc_col_idx = find_column_by_header_in_range(
                ws, desc_header, DESC_START_COL, DESC_END_COL, header_row
            )
            if desc_col_idx:
                desc_val = ws.cell(row=row, column=desc_col_idx).value or ""
        deps.append({"equipo": equipo, "codigo": flag_up, "descripcion": desc_val})
    return deps


def summarize_project_by_row(row: int) -> Dict[str, Any]:
    wb = load_workbook()
    ws = get_ws_proyectos(wb)
    project_data = read_project_row(ws, row)
    deps = get_project_dependencies(ws, row)

    total_dep, total_L, total_P, cub = compute_dep_aggregates(deps)

    return {
        "row": row,
        "project": project_data,
        "dependencias": deps,
        "total_dep": total_dep,
        "total_L": total_L,
        "total_P": total_P,
        "cubrimiento": cub,
    }
